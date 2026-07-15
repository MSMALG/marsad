

import json
import re
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUT_DIR = Path("/mnt/user-data/uploads")
OUTPUT_DIR = Path("/mnt/user-data/outputs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

HH_XLSX = INPUT_DIR / "Publication_tables_of_Household_Income_and_Consumption_Expenditure_Survey_2023_2__1_.xlsx"

report_lines = []  # collects markdown notes as we go, for business_rules_report.md


def log(msg):
    print(msg)
    report_lines.append(msg)


# ---------------------------------------------------------------------------
# 1. POPULATION — age/gender/nationality distribution
# ---------------------------------------------------------------------------
def parse_population():
    df = pd.read_csv(INPUT_DIR / "Population.csv", header=[0, 1])
    df.columns = [
        "age_group",
        "saudi_male", "saudi_female", "saudi_total",
        "nonsaudi_male", "nonsaudi_female", "nonsaudi_total",
        "total_male", "total_female", "total_total",
    ]
    df = df.iloc[1:].copy()  # drop the duplicated header row baked into the CSV
    df["age_group"] = df["age_group"].str.strip()
    df = df[df["age_group"].str.lower() != "total"]

    for col in df.columns[1:]:
        df[col] = (
            df[col].astype(str).str.replace(",", "", regex=False).astype(float)
        )

    total_pop = df["total_total"].sum()
    df["age_group"] = df["age_group"].str.replace("^0(\\d)-", r"\1-", regex=True)

    age_distribution = {
        row.age_group: {
            "saudi_male": row.saudi_male,
            "saudi_female": row.saudi_female,
            "nonsaudi_male": row.nonsaudi_male,
            "nonsaudi_female": row.nonsaudi_female,
            "share_of_population": round(row.total_total / total_pop, 6),
        }
        for row in df.itertuples()
    }

    saudi_total = df["saudi_total"].sum()
    nonsaudi_total = df["nonsaudi_total"].sum()
    nationality_split = {
        "saudi_share": round(saudi_total / total_pop, 4),
        "nonsaudi_share": round(nonsaudi_total / total_pop, 4),
    }

    log(f"## Population\nParsed {len(df)} age bands. Total population = {int(total_pop):,}. "
        f"Saudi share = {nationality_split['saudi_share']:.1%}, "
        f"Non-Saudi share = {nationality_split['nonsaudi_share']:.1%}.\n")

    return {
        "source": "GASTAT Population Dataset",
        "total_population": int(total_pop),
        "nationality_split": nationality_split,
        "age_group_distribution": age_distribution,
    }


# ---------------------------------------------------------------------------
# 2. WAGES — average monthly wage by nationality/gender, latest quarter
# ---------------------------------------------------------------------------
def parse_wages():
    # Row 0 is a bilingual header ("Nationality, Year / Quarter, Unnamed..."),
    # row 1 holds the real quarter labels ("Gender, 2016 / Q2, ..."), data starts row 2.
    raw = pd.read_csv(INPUT_DIR / "Wages_of_Employees_by_gender_and_nationality.csv", header=None)
    quarter_labels = raw.iloc[1, 1:].tolist()
    data = raw.iloc[2:].copy()
    data.columns = ["group"] + quarter_labels
    latest_q = quarter_labels[-1]

    # Rows alternate: Saudi, Male, Female, Non Saudi, Male, Female
    wages = {
        str(r["group"]).strip(): round(float(r[latest_q]), 2)
        for _, r in data.iterrows()
    }

    log(f"## Wages\nLatest quarter in data: **{latest_q}**. "
        f"Average monthly wage (SAR): Saudi = {wages.get('Saudi')}, "
        f"Non Saudi = {wages.get('Non Saudi')}.\n")

    return {
        "source": "GASTAT Wages of Employees by Gender and Nationality",
        "latest_quarter": latest_q,
        "average_monthly_wage_sar": wages,
    }


# ---------------------------------------------------------------------------
# 3. CONSUMER PRICE INDEX — regional cost-of-living adjustment factors
# ---------------------------------------------------------------------------
def _clean_cpi_dataframe(df):
    df.columns = [c.strip().rstrip(";").strip() for c in df.columns]
    keep = ["السنة", "التاريخ", "الشهر", "المنطقة الإدارية", "المستوى",
            "البند الإنجليزي", "الرقم القياسي", "الأهمية النسبية"]
    df = df[[c for c in keep if c in df.columns]].copy()
    df["الرقم القياسي"] = pd.to_numeric(
        df["الرقم القياسي"].astype(str).str.replace(";", "", regex=False), errors="coerce"
    )
    return df.dropna(subset=["الرقم القياسي"])


def parse_cpi():
    files = {
        "2025-11": ("Consumer_Price_Index_csv__2_.csv", dict(skiprows=1)),
        "2025-12": ("Consumer_Price_Index_csv__1_.csv", dict()),
        "2026-01": ("Consumer_Price_Index_csv.csv", dict(engine="python", on_bad_lines="skip")),
    }
    frames = {}
    for month, (fname, kwargs) in files.items():
        try:
            raw = pd.read_csv(INPUT_DIR / fname, **kwargs)
            frames[month] = _clean_cpi_dataframe(raw)
        except Exception as e:
            log(f"  - WARNING: could not parse {fname}: {e}")

    latest_month = sorted(frames.keys())[-1]
    latest = frames[latest_month]

    # General Index (level 0, item code 0) per region -> regional cost-of-living factor
    general = latest[(latest["المستوى"] == 0)]
    national_index = general.loc[general["المنطقة الإدارية"] == "المملكة", "الرقم القياسي"]
    national_index = float(national_index.iloc[0]) if len(national_index) else general["الرقم القياسي"].mean()

    regional_cpi = {}
    for _, row in general.iterrows():
        regional_cpi[row["المنطقة الإدارية"]] = {
            "general_index": round(float(row["الرقم القياسي"]), 4),
            "cost_of_living_factor": round(float(row["الرقم القياسي"]) / national_index, 4),
        }

    # Main COICOP-style divisions at level 1 (national only) for category-level indices
    level1_national = latest[
        (latest["المستوى"] == 1) & (latest["المنطقة الإدارية"] == "المملكة")
    ]
    category_index = {
        str(r["البند الإنجليزي"]).strip(): round(float(r["الرقم القياسي"]), 4)
        for _, r in level1_national.iterrows()
    }

    log(f"## Consumer Price Index\nMonths parsed: {sorted(frames.keys())}. Using **{latest_month}** as latest.\n"
        f"Note: the Jan 2026 export ({files['2026-01'][0]}) has malformed quoting in the source file "
        f"(unescaped commas inside item names) and was recovered with `on_bad_lines='skip'`, losing "
        f"~290 of 1540 rows. December 2025 remains the cleanest full month if more precision is needed.\n"
        f"Regions found: {list(regional_cpi.keys())}.\n")

    return {
        "source": "GASTAT Consumer Price Index",
        "latest_month": latest_month,
        "national_general_index": round(national_index, 4),
        "regional_cost_of_living": regional_cpi,
        "category_index_national": category_index,
    }


# ---------------------------------------------------------------------------
# 4. HOUSEHOLD INCOME & CONSUMPTION EXPENDITURE SURVEY 2023 (the core dataset)
# ---------------------------------------------------------------------------
def _sheet_rows(wb, sheet_name):
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if vals:
            rows.append(vals)
    return rows


def _clean_label(label):
    """Take the English half of a bilingual 'Arabic\\nEnglish' label."""
    if not isinstance(label, str):
        return str(label)
    parts = label.split("\n")
    eng = parts[-1] if len(parts) > 1 else parts[0]
    return re.sub(r"\s+", " ", eng).strip()


def _parse_saudi_nonsaudi_table(rows):
    """Tables shaped: [category, saudi, nonsaudi, total] starting after the header rows."""
    data = {}
    for r in rows:
        if len(r) >= 4 and isinstance(r[1], (int, float)):
            label = _clean_label(r[0])
            if label.lower().startswith("total"):
                continue
            data[label] = {"saudi": r[1], "non_saudi": r[2], "total": r[3]}
    return data


def _parse_region_table(rows, regions):
    """Tables shaped: [category, val_region1, val_region2, ...] matching `regions` order."""
    data = {}
    for r in rows:
        if len(r) >= len(regions) + 1 and all(isinstance(v, (int, float)) for v in r[1:1 + len(regions)]):
            label = _clean_label(r[0])
            data[label] = {reg: r[i + 1] for i, reg in enumerate(regions)}
    return data


def parse_household_survey():
    wb = load_workbook(HH_XLSX, read_only=True, data_only=True)
    result = {"source": "GASTAT Household Income and Consumption Expenditure Survey 2023"}

    # --- 1-1 Main indicators ---
    rows = _sheet_rows(wb, "1-1")
    result["main_indicators"] = _parse_saudi_nonsaudi_table(rows)

    # --- Income breakdowns (by nationality of household head) ---
    income = {}
    income["by_region"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "1-2"))
    income["by_age_group"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "3-2"))
    income["by_education"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "4-2"))
    income["by_marital_status"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "5-2"))
    income["by_dwelling_type"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "6-2"))
    income["by_ownership_type"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "7-2 "))
    income["by_household_size"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "8-2"))
    income["sources_pct"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "9-2"))
    result["disposable_income_sar_per_month"] = income

    # --- Expenditure breakdowns (Consumption Expenditure, by nationality of household head) ---
    expenditure = {}
    expenditure["by_region"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "1-3"))
    expenditure["by_age_group"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "3-3"))
    expenditure["by_education"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "4-3 "))
    expenditure["by_marital_status"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "5-3"))
    expenditure["by_dwelling_type"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "6-3"))
    expenditure["by_ownership_type"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "7-3"))
    expenditure["by_household_size"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "8-3"))
    result["consumption_expenditure_sar_per_month"] = expenditure

    # --- COICOP category breakdown (this drives Food/Housing/Transport/... expense columns) ---
    coicop = {}
    coicop["avg_by_nationality"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "11-4"))
    coicop["pct_by_nationality"] = _parse_saudi_nonsaudi_table(_sheet_rows(wb, "12-4"))

    regions = ["Riyadh", "Makkah", "Madinah", "Qassim", "Eastern Province", "Asir", "Tabuk",
               "Hail", "Northern Border", "Jazan", "Najran", "AL - Baha", "AL - Jouf"]
    coicop["avg_by_region"] = _parse_region_table(_sheet_rows(wb, "15-4"), regions)
    coicop["pct_by_region"] = _parse_region_table(_sheet_rows(wb, "16-4"), regions)
    result["coicop_final_monetary_consumption"] = coicop

    # --- Health expenditure burden ---
    health_rows = _sheet_rows(wb, "17-4")
    health = {}
    for r in health_rows:
        if len(r) == 2 and isinstance(r[1], (int, float)):
            health[_clean_label(r[0])] = round(r[1], 4)
    result["health_expenditure_burden_pct_of_households"] = health

    log(f"## Household Income & Consumption Expenditure Survey 2023\n"
        f"Extracted {len(income)} income breakdowns, {len(expenditure)} expenditure breakdowns, "
        f"and full COICOP (12-division) spending category tables by nationality and by region.\n")

    return result


# ---------------------------------------------------------------------------
# 5. COICOP -> MERSAD EXPENSE-COLUMN MAPPING
#    Maps the 12 official COICOP divisions onto the 5 expense columns from the
#    project spec (Food / Housing / Transportation / Entertainment / Healthcare)
#    plus a catch-all "Other" bucket for the remaining divisions.
# ---------------------------------------------------------------------------
COICOP_TO_MERSAD = {
    "Food_Expense": ["Food and beverages", "Tobacco"],
    "Housing_Expense": ["Housing, water, electricity, gas and other fuels"],
    "Transportation_Expense": ["Transport"],
    "Entertainment_Expense": ["Recreation, sport and culture", "Restaurants and accommodation services"],
    "Healthcare_Expense": ["Health"],
    "Other_Expense": [
        "Clothing and footwear",
        "Furnishings, household equipment and routine household maintenance",
        "Information and communication",
        "Education services",
        "Insurance and financial services",
        "Personal care, social protection and miscellaneous goods and services",
    ],
}


def build_expense_category_shares(household_survey):
    """Collapse the 12 COICOP % divisions into the 6 Mersad expense buckets,
    both nationally (by nationality) and by region."""

    def _label_lookup(pct_table):
        # normalise keys (strip trailing punctuation/whitespace differences)
        return {k.strip(): v for k, v in pct_table.items()}

    pct_nat = _label_lookup(household_survey["coicop_final_monetary_consumption"]["pct_by_nationality"])
    pct_reg = household_survey["coicop_final_monetary_consumption"]["pct_by_region"]

    def collapse_nat(pct_table, key):
        out = {}
        for bucket, coicop_items in COICOP_TO_MERSAD.items():
            total = 0.0
            for item in coicop_items:
                match = next((v for k, v in pct_table.items() if k.strip().startswith(item[:15])), None)
                if match:
                    total += match[key]
            out[bucket] = round(total, 4)
        return out

    shares_by_nationality = {
        "saudi": collapse_nat(pct_nat, "saudi"),
        "non_saudi": collapse_nat(pct_nat, "non_saudi"),
        "total": collapse_nat(pct_nat, "total"),
    }

    shares_by_region = {}
    for region in ["Riyadh", "Makkah", "Madinah", "Qassim", "Eastern Province", "Asir", "Tabuk",
                    "Hail", "Northern Border", "Jazan", "Najran", "AL - Baha", "AL - Jouf"]:
        bucket_shares = {}
        for bucket, coicop_items in COICOP_TO_MERSAD.items():
            total = 0.0
            for item in coicop_items:
                match = next((v for k, v in pct_reg.items() if k.strip().startswith(item[:15])), None)
                if match:
                    total += match.get(region, 0)
            bucket_shares[bucket] = round(total, 4)
        shares_by_region[region] = bucket_shares

    log("## Mersad Expense-Category Mapping\n"
        "Collapsed 12 official COICOP divisions into 6 Mersad expense buckets "
        "(Food, Housing, Transportation, Entertainment, Healthcare, Other). "
        "Shares are computed both by nationality and by region so the synthetic "
        "generator can allocate a household's total spending realistically.\n")

    return {
        "mapping_definition": COICOP_TO_MERSAD,
        "shares_by_nationality": shares_by_nationality,
        "shares_by_region": shares_by_region,
    }


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
REGION_AR_TO_EN = {
    "المملكة": "Kingdom (National)",
    "الرياض": "Riyadh",
    "مكة المكرمة": "Makkah",
    "المدينة المنورة": "Madinah",
    "القصيم": "Qassim",
    "المنطقة الشرقية": "Eastern Province",
    "عسير": "Asir",
    "تبوك": "Tabuk",
    "حائل": "Hail",
    "الحدود الشمالية": "Northern Border",
    "جازان": "Jazan",
    "نجران": "Najran",
    "الباحة": "AL - Baha",
    "الجوف": "AL - Jouf",
}


def main():
    log("# Mersad AI — Business Rules Extraction Report\n")

    population = parse_population()
    wages = parse_wages()
    cpi = parse_cpi()
    cpi["region_name_bridge_ar_to_en"] = REGION_AR_TO_EN
    cpi["regional_cost_of_living_en"] = {
        REGION_AR_TO_EN.get(ar, ar): v for ar, v in cpi["regional_cost_of_living"].items()
    }
    household_survey = parse_household_survey()
    expense_mapping = build_expense_category_shares(household_survey)

    business_rules = {
        "metadata": {
            "project": "Mersad AI",
            "description": "Statistical distributions and business rules derived from official "
                            "Saudi government datasets, used to generate a realistic synthetic "
                            "personal-finance dataset (100k-300k rows) for ML model training.",
            "sources": [
                "Saudi Open Data Platform",
                "General Authority for Statistics (GASTAT)",
            ],
        },
        "population": population,
        "wages": wages,
        "consumer_price_index": cpi,
        "household_income_and_expenditure": household_survey,
        "mersad_expense_category_mapping": expense_mapping,
    }

    out_json = OUTPUT_DIR / "business_rules.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(business_rules, f, ensure_ascii=False, indent=2)

    out_md = OUTPUT_DIR / "business_rules_report.md"
    with open(out_md, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))

    print(f"\nSaved: {out_json}")
    print(f"Saved: {out_md}")


if __name__ == "__main__":
    main()
